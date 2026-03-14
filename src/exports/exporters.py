"""Export module for generating Power BI / Tableau-ready files.

Produces CSV and formatted Excel exports from the insurance claims
data warehouse.  All exports are derived from the star-schema views
defined in ``sql/views.sql`` so that downstream BI tools receive
fully denormalised, analysis-ready datasets.

Usage
-----
    python -m src.exports.exporters          # CLI
    # or
    from src.exports.exporters import export_all
    paths = export_all()
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from sqlalchemy import text

from config import DB_URL, EXPORTS_DIR
from src.database.connection import get_engine


# ── Constants ────────────────────────────────────────────────────────────

_TIMESTAMP_FMT = "%Y%m%d_%H%M%S"

# SQL that replicates v_claims_summary inline so the export works even if
# the view has not been materialised.
_CLAIMS_SUMMARY_SQL = """
SELECT
    f.claim_id,
    f.incident_date,
    f.report_date,
    f.close_date,
    f.claim_type,
    f.procedure_category,
    f.severity_level,
    f.status,
    f.paid_amount,
    f.incurred_amount,
    f.reserved_amount,
    f.days_to_close,
    f.days_to_report,
    f.patient_age_band,
    f.patient_risk_segment,
    f.repeat_event_flag,
    f.litigation_flag,
    f.accident_year,
    f.development_year,

    p.provider_id,
    p.provider_name,
    p.specialty,
    p.department,
    p.risk_tier,

    r.state_code,
    r.state_name,
    r.region,
    r.urban_rural,

    d.diagnosis_code,
    d.diagnosis_category,
    d.severity_weight,

    rc.root_cause_code,
    rc.root_cause_category,
    rc.preventability,

    ROUND(f.paid_amount / NULLIF(f.incurred_amount, 0), 4) AS loss_ratio,
    ROUND(f.reserved_amount - f.paid_amount, 2)            AS reserve_adequacy,
    SUBSTR(f.incident_date, 1, 7)                          AS incident_month,
    CASE WHEN f.severity_level >= 4 THEN 1 ELSE 0 END     AS is_high_severity,
    CASE WHEN f.days_to_close > 365  THEN 1 ELSE 0 END    AS is_long_tail

FROM fact_claim f
JOIN dim_provider   p  ON f.provider_key   = p.provider_key
JOIN dim_region     r  ON f.region_key     = r.region_key
JOIN dim_diagnosis  d  ON f.diagnosis_key  = d.diagnosis_key
JOIN dim_root_cause rc ON f.root_cause_key = rc.root_cause_key
"""


# ── Helpers ──────────────────────────────────────────────────────────────

def _ensure_dir(path: Path) -> Path:
    """Create *path* (and parents) if it does not already exist."""
    path.mkdir(parents=True, exist_ok=True)
    return path


def _load_claims_df(engine) -> pd.DataFrame:
    """Load the full claims summary from the warehouse.

    Tries the ``v_claims_summary`` view first; falls back to an inline
    join if the view does not exist.
    """
    with engine.connect() as conn:
        try:
            df = pd.read_sql(text("SELECT * FROM v_claims_summary"), conn)
        except Exception:
            df = pd.read_sql(text(_CLAIMS_SUMMARY_SQL), conn)
    return df


# ── CSV Exporters ────────────────────────────────────────────────────────

def export_claims_csv(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write the full claims fact table (with dimensions joined) to CSV.

    Parameters
    ----------
    df : pd.DataFrame
        Claims summary DataFrame (one row per claim).
    output_dir : Path
        Directory where the CSV will be created.

    Returns
    -------
    Path
        Absolute path to the written file.
    """
    _ensure_dir(output_dir)
    path = output_dir / "claims_detail.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  [CSV]  {path}  ({len(df):,} rows)")
    return path


def export_loss_triangle_csv(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write the loss-development triangle as a pivot-table CSV.

    Rows are accident years, columns are development years, and cell
    values are cumulative paid amounts.

    Parameters
    ----------
    df : pd.DataFrame
        Claims summary DataFrame with *accident_year*, *development_year*,
        and *paid_amount* columns.
    output_dir : Path
        Destination directory.

    Returns
    -------
    Path
        Absolute path to the written file.
    """
    _ensure_dir(output_dir)

    df = df.copy()
    df["accident_year"] = pd.to_numeric(df["accident_year"], errors="coerce")
    df["development_year"] = pd.to_numeric(df["development_year"], errors="coerce")
    df["paid_amount"] = pd.to_numeric(df["paid_amount"], errors="coerce").fillna(0)
    df = df.dropna(subset=["accident_year", "development_year"])

    # Incremental amounts
    incremental = (
        df.groupby(["accident_year", "development_year"])["paid_amount"]
        .sum()
        .reset_index()
    )

    # Pivot and cumulate across development years
    pivot = incremental.pivot_table(
        index="accident_year",
        columns="development_year",
        values="paid_amount",
        aggfunc="sum",
        fill_value=0,
    )
    pivot = pivot.sort_index(axis=0).sort_index(axis=1)
    cumulative = pivot.cumsum(axis=1).round(2)

    # Rename columns for readability
    cumulative.columns = [f"dev_{int(c)}" for c in cumulative.columns]
    cumulative.index = cumulative.index.astype(int)
    cumulative.index.name = "accident_year"

    path = output_dir / "loss_triangle.csv"
    cumulative.to_csv(path, encoding="utf-8-sig")
    print(f"  [CSV]  {path}  ({len(cumulative):,} rows)")
    return path


def export_monthly_metrics_csv(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write monthly time-series metrics to CSV.

    Aggregates claims by incident month and computes count, total paid,
    total incurred, average severity, average days to close, litigation
    count, and reopen count.

    Parameters
    ----------
    df : pd.DataFrame
        Claims summary DataFrame.
    output_dir : Path
        Destination directory.

    Returns
    -------
    Path
        Absolute path to the written file.
    """
    _ensure_dir(output_dir)

    df = df.copy()
    df["incident_date"] = pd.to_datetime(df["incident_date"], errors="coerce")
    df["paid_amount"] = pd.to_numeric(df["paid_amount"], errors="coerce").fillna(0)
    df["incurred_amount"] = pd.to_numeric(df["incurred_amount"], errors="coerce").fillna(0)
    df["severity_level"] = pd.to_numeric(df["severity_level"], errors="coerce")
    df["days_to_close"] = pd.to_numeric(df["days_to_close"], errors="coerce")
    df["litigation_flag"] = pd.to_numeric(df["litigation_flag"], errors="coerce").fillna(0)

    df["month"] = df["incident_date"].dt.to_period("M").astype(str)

    monthly = df.groupby("month", sort=True).agg(
        claim_count=("claim_id", "count"),
        total_paid=("paid_amount", "sum"),
        total_incurred=("incurred_amount", "sum"),
        avg_severity=("severity_level", "mean"),
        avg_days_to_close=("days_to_close", "mean"),
        litigation_count=("litigation_flag", "sum"),
        reopen_count=("status", lambda s: (s == "reopened").sum()),
    ).reset_index()

    # Round monetary and average columns
    monthly["total_paid"] = monthly["total_paid"].round(2)
    monthly["total_incurred"] = monthly["total_incurred"].round(2)
    monthly["avg_severity"] = monthly["avg_severity"].round(2)
    monthly["avg_days_to_close"] = monthly["avg_days_to_close"].round(1)
    monthly["litigation_count"] = monthly["litigation_count"].astype(int)
    monthly["reopen_count"] = monthly["reopen_count"].astype(int)

    path = output_dir / "monthly_metrics.csv"
    monthly.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  [CSV]  {path}  ({len(monthly):,} rows)")
    return path


def export_provider_scorecard_csv(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write per-provider risk scorecards to CSV.

    Columns: provider_id, provider_name, specialty, department, risk_tier,
    claim_count, avg_severity, avg_paid, total_paid, litigation_rate,
    avg_days_to_close, avg_days_to_report, reopen_count, repeat_event_count.

    Parameters
    ----------
    df : pd.DataFrame
        Claims summary DataFrame.
    output_dir : Path
        Destination directory.

    Returns
    -------
    Path
        Absolute path to the written file.
    """
    _ensure_dir(output_dir)

    df = df.copy()
    df["paid_amount"] = pd.to_numeric(df["paid_amount"], errors="coerce").fillna(0)
    df["severity_level"] = pd.to_numeric(df["severity_level"], errors="coerce")
    df["days_to_close"] = pd.to_numeric(df["days_to_close"], errors="coerce")
    df["days_to_report"] = pd.to_numeric(df["days_to_report"], errors="coerce")
    df["litigation_flag"] = pd.to_numeric(df["litigation_flag"], errors="coerce").fillna(0)
    df["repeat_event_flag"] = pd.to_numeric(df["repeat_event_flag"], errors="coerce").fillna(0)

    group_cols = ["provider_id", "provider_name", "specialty", "department", "risk_tier"]
    scorecard = df.groupby(group_cols, sort=True).agg(
        claim_count=("claim_id", "count"),
        avg_severity=("severity_level", "mean"),
        avg_paid=("paid_amount", "mean"),
        total_paid=("paid_amount", "sum"),
        litigation_rate=("litigation_flag", "mean"),
        avg_days_to_close=("days_to_close", "mean"),
        avg_days_to_report=("days_to_report", "mean"),
        reopen_count=("status", lambda s: (s == "reopened").sum()),
        repeat_event_count=("repeat_event_flag", "sum"),
    ).reset_index()

    # Rounding
    scorecard["avg_severity"] = scorecard["avg_severity"].round(2)
    scorecard["avg_paid"] = scorecard["avg_paid"].round(2)
    scorecard["total_paid"] = scorecard["total_paid"].round(2)
    scorecard["litigation_rate"] = scorecard["litigation_rate"].round(4)
    scorecard["avg_days_to_close"] = scorecard["avg_days_to_close"].round(1)
    scorecard["avg_days_to_report"] = scorecard["avg_days_to_report"].round(1)
    scorecard["reopen_count"] = scorecard["reopen_count"].astype(int)
    scorecard["repeat_event_count"] = scorecard["repeat_event_count"].astype(int)

    # Sort by total_paid descending for quick identification of high-cost providers
    scorecard = scorecard.sort_values("total_paid", ascending=False)

    path = output_dir / "provider_scorecard.csv"
    scorecard.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  [CSV]  {path}  ({len(scorecard):,} rows)")
    return path


# ── Excel Exporter ───────────────────────────────────────────────────────

def export_claims_excel(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write a multi-sheet Excel workbook with formatted analytics.

    Sheets
    ------
    1. **Claims Detail** -- full denormalised claims data.
    2. **Monthly Summary** -- monthly aggregated metrics.
    3. **Severity Analysis** -- severity distribution by claim type.
    4. **Provider Scorecard** -- per-provider risk metrics.
    5. **Regional Summary** -- per-region metrics.

    Formatting
    ----------
    - Bold headers with auto-filter enabled.
    - Column widths auto-fitted to content.
    - Monetary columns formatted as ``#,##0.00``.
    - Percentage / rate columns formatted as ``0.00%``.

    Parameters
    ----------
    df : pd.DataFrame
        Claims summary DataFrame.
    output_dir : Path
        Destination directory.

    Returns
    -------
    Path
        Absolute path to the written file.
    """
    from openpyxl.styles import Font, numbers
    from openpyxl.utils import get_column_letter

    _ensure_dir(output_dir)
    path = output_dir / "claims_dashboard.xlsx"

    df = df.copy()
    # Ensure numeric types
    for col in ("paid_amount", "incurred_amount", "reserved_amount",
                "severity_level", "days_to_close", "days_to_report",
                "litigation_flag", "repeat_event_flag", "severity_weight",
                "loss_ratio", "reserve_adequacy", "is_high_severity",
                "is_long_tail"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── Build sheets ──────────────────────────────────────────────────────

    # 1) Claims Detail -- direct export of the summary
    detail_df = df

    # 2) Monthly Summary
    df["incident_date"] = pd.to_datetime(df["incident_date"], errors="coerce")
    df["_month"] = df["incident_date"].dt.to_period("M").astype(str)

    monthly_df = df.groupby("_month", sort=True).agg(
        claim_count=("claim_id", "count"),
        total_paid=("paid_amount", "sum"),
        total_incurred=("incurred_amount", "sum"),
        total_reserved=("reserved_amount", "sum"),
        avg_severity=("severity_level", "mean"),
        avg_days_to_close=("days_to_close", "mean"),
        avg_days_to_report=("days_to_report", "mean"),
        litigation_count=("litigation_flag", "sum"),
        high_severity_count=("is_high_severity", "sum"),
    ).reset_index().rename(columns={"_month": "month"})

    for c in ("total_paid", "total_incurred", "total_reserved"):
        monthly_df[c] = monthly_df[c].round(2)
    monthly_df["avg_severity"] = monthly_df["avg_severity"].round(2)
    monthly_df["avg_days_to_close"] = monthly_df["avg_days_to_close"].round(1)
    monthly_df["avg_days_to_report"] = monthly_df["avg_days_to_report"].round(1)
    monthly_df["litigation_count"] = monthly_df["litigation_count"].astype(int)
    monthly_df["high_severity_count"] = monthly_df["high_severity_count"].astype(int)

    # 3) Severity Analysis -- severity distribution by claim_type
    severity_df = (
        df.groupby(["claim_type", "severity_level"])
        .agg(
            claim_count=("claim_id", "count"),
            total_paid=("paid_amount", "sum"),
            avg_paid=("paid_amount", "mean"),
        )
        .reset_index()
    )
    severity_df["total_paid"] = severity_df["total_paid"].round(2)
    severity_df["avg_paid"] = severity_df["avg_paid"].round(2)

    # 4) Provider Scorecard
    provider_df = df.groupby(
        ["provider_id", "provider_name", "specialty", "department", "risk_tier"],
        sort=True,
    ).agg(
        claim_count=("claim_id", "count"),
        avg_severity=("severity_level", "mean"),
        avg_paid=("paid_amount", "mean"),
        total_paid=("paid_amount", "sum"),
        litigation_rate=("litigation_flag", "mean"),
        avg_days_to_close=("days_to_close", "mean"),
        avg_days_to_report=("days_to_report", "mean"),
        reopen_count=("status", lambda s: (s == "reopened").sum()),
        repeat_event_count=("repeat_event_flag", "sum"),
    ).reset_index().sort_values("total_paid", ascending=False)

    provider_df["avg_severity"] = provider_df["avg_severity"].round(2)
    provider_df["avg_paid"] = provider_df["avg_paid"].round(2)
    provider_df["total_paid"] = provider_df["total_paid"].round(2)
    provider_df["litigation_rate"] = provider_df["litigation_rate"].round(4)
    provider_df["avg_days_to_close"] = provider_df["avg_days_to_close"].round(1)
    provider_df["avg_days_to_report"] = provider_df["avg_days_to_report"].round(1)

    # 5) Regional Summary
    regional_df = df.groupby(
        ["state_code", "state_name", "region", "urban_rural"],
        sort=True,
    ).agg(
        claim_count=("claim_id", "count"),
        total_paid=("paid_amount", "sum"),
        total_incurred=("incurred_amount", "sum"),
        avg_severity=("severity_level", "mean"),
        avg_days_to_close=("days_to_close", "mean"),
        litigation_rate=("litigation_flag", "mean"),
        high_severity_pct=("is_high_severity", "mean"),
    ).reset_index().sort_values("total_paid", ascending=False)

    regional_df["total_paid"] = regional_df["total_paid"].round(2)
    regional_df["total_incurred"] = regional_df["total_incurred"].round(2)
    regional_df["avg_severity"] = regional_df["avg_severity"].round(2)
    regional_df["avg_days_to_close"] = regional_df["avg_days_to_close"].round(1)
    regional_df["litigation_rate"] = regional_df["litigation_rate"].round(4)
    regional_df["high_severity_pct"] = regional_df["high_severity_pct"].round(4)

    # ── Write to Excel ────────────────────────────────────────────────────

    sheets = {
        "Claims Detail": detail_df,
        "Monthly Summary": monthly_df,
        "Severity Analysis": severity_df,
        "Provider Scorecard": provider_df,
        "Regional Summary": regional_df,
    }

    # Columns that should receive monetary number formatting
    _money_cols = {
        "paid_amount", "incurred_amount", "reserved_amount",
        "total_paid", "total_incurred", "total_reserved",
        "avg_paid", "reserve_adequacy",
    }
    # Columns that should receive percentage formatting
    _pct_cols = {
        "litigation_rate", "loss_ratio", "high_severity_pct",
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Post-write formatting via openpyxl workbook handle
        wb = writer.book
        bold_font = Font(bold=True)

        for sheet_name, sheet_df in sheets.items():
            ws = wb[sheet_name]
            n_cols = len(sheet_df.columns)
            n_rows = len(sheet_df)

            # Bold header row
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=1, column=col_idx).font = bold_font

            # Auto-filter over entire data range
            last_col_letter = get_column_letter(n_cols)
            ws.auto_filter.ref = f"A1:{last_col_letter}{n_rows + 1}"

            # Column-width auto-fit and number formatting
            for col_idx in range(1, n_cols + 1):
                col_name = sheet_df.columns[col_idx - 1]
                col_letter = get_column_letter(col_idx)

                # Determine max width from header + data
                max_width = len(str(col_name))
                for row_idx in range(2, min(n_rows + 2, 102)):  # sample first 100 rows
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val is not None:
                        max_width = max(max_width, len(str(cell_val)))
                # Cap and pad
                ws.column_dimensions[col_letter].width = min(max_width + 3, 50)

                # Apply number formatting to data cells
                if col_name in _money_cols:
                    for row_idx in range(2, n_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = (
                            numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # #,##0.00
                        )
                elif col_name in _pct_cols:
                    for row_idx in range(2, n_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

            # Freeze the header row
            ws.freeze_panes = "A2"

    print(f"  [XLSX] {path}  (5 sheets, {len(detail_df):,} detail rows)")
    return path


# ── Main Orchestrator ────────────────────────────────────────────────────

def export_all(db_url: Optional[str] = None) -> Dict[str, Path]:
    """Generate all BI-ready exports from the claims warehouse.

    Parameters
    ----------
    db_url : str or None
        SQLAlchemy database URL.  Defaults to ``config.DB_URL``.

    Returns
    -------
    dict
        Mapping of export name to absolute file path, e.g.::

            {
                "claims_csv":           Path("data/exports/claims_detail.csv"),
                "claims_excel":         Path("data/exports/claims_dashboard.xlsx"),
                "loss_triangle_csv":    Path("data/exports/loss_triangle.csv"),
                "monthly_metrics_csv":  Path("data/exports/monthly_metrics.csv"),
                "provider_scorecard_csv": Path("data/exports/provider_scorecard.csv"),
            }
    """
    url = db_url or DB_URL
    output_dir = EXPORTS_DIR

    print(f"\n{'=' * 60}")
    print("  EXPORT GENERATOR")
    print(f"{'=' * 60}")
    print(f"\n  Database : {url}")
    print(f"  Output   : {output_dir}\n")

    engine = get_engine(url)
    df = _load_claims_df(engine)
    print(f"  Loaded {len(df):,} claims from warehouse\n")

    if df.empty:
        print("  WARNING: No claims data found. Skipping exports.")
        return {}

    paths: Dict[str, Path] = {}

    paths["claims_csv"] = export_claims_csv(df, output_dir)
    paths["claims_excel"] = export_claims_excel(df, output_dir)
    paths["loss_triangle_csv"] = export_loss_triangle_csv(df, output_dir)
    paths["monthly_metrics_csv"] = export_monthly_metrics_csv(df, output_dir)
    paths["provider_scorecard_csv"] = export_provider_scorecard_csv(df, output_dir)

    print(f"\n  {len(paths)} exports written to {output_dir}")
    print(f"{'=' * 60}\n")

    return paths


# ── CLI entry point ──────────────────────────────────────────────────────

if __name__ == "__main__":
    export_all()
